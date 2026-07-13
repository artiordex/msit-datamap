from __future__ import annotations

import argparse
import json
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv

load_dotenv()


CATEGORY_BY_SHEET_KEY = {
    "공공데이터민간활용": "공공데이터 민간 활용 사례",
    "AI도입활용": "AI 도입 및 업무 활용 사례",
    "데이터분석활용": "데이터 분석 활용 사례",
}

INSTITUTION_HEADERS = ("기관명", "기관", "기관명(약칭)", "기관명칭", "대상기관")
KEYWORD_HEADERS = ("검색어", "키워드", "질문", "검색질의", "비고")


@dataclass
class SearchItem:
    sheet: str
    category: str
    row_number: int
    institution: str
    query: str
    row: dict[str, Any]


def load_project_env(project_dir: Path) -> None:
    load_dotenv(project_dir / ".env")


def resolve_input(project_dir: Path, input_file: str) -> Path:
    path = Path(input_file)
    if not path.is_absolute():
        path = project_dir / path
    return path


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_filename(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣_.-]+", "_", value).strip("_") or "result"


def sheet_key(sheet_name: str) -> str:
    return re.sub(r"\s+", "", sheet_name)


def category_for_sheet(sheet_name: str) -> str:
    return CATEGORY_BY_SHEET_KEY.get(sheet_key(sheet_name), sheet_name)


def find_header_index(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    for candidate in candidates:
        for index, header in enumerate(headers):
            if candidate in header:
                return index
    return None


def row_to_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for index, value in enumerate(values):
        key = headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}"
        result[key] = value
    return result


def build_query(sheet_name: str, headers: list[str], values: tuple[Any, ...]) -> tuple[str, str]:
    category = category_for_sheet(sheet_name)
    institution_index = find_header_index(headers, INSTITUTION_HEADERS)
    keyword_index = find_header_index(headers, KEYWORD_HEADERS)

    institution = normalize(values[institution_index]) if institution_index is not None and institution_index < len(values) else ""
    keyword = normalize(values[keyword_index]) if keyword_index is not None and keyword_index < len(values) else ""

    parts = [institution, category, keyword, "국내 공공기관 활용 사례"]
    query = " ".join(part for part in parts if part)
    return institution, query


def collect_search_items(workbook_path: Path, max_rows: int) -> list[SearchItem]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    items: list[SearchItem] = []
    sheet_names = workbook.sheetnames
    selected_sheets = [name for name in sheet_names if sheet_key(name) in CATEGORY_BY_SHEET_KEY] or sheet_names

    for sheet_name in selected_sheets:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            continue
        headers = [normalize(value) for value in header_row]

        for row_offset, values in enumerate(rows, start=2):
            if len(items) >= max_rows:
                workbook.close()
                return items
            if not any(normalize(value) for value in values):
                continue
            institution, query = build_query(sheet_name, headers, values)
            if not query:
                continue
            items.append(
                SearchItem(
                    sheet=sheet_name,
                    category=category_for_sheet(sheet_name),
                    row_number=row_offset,
                    institution=institution,
                    query=query,
                    row=row_to_dict(headers, values),
                )
            )

    workbook.close()
    return items


def extract_sources(response_dict: dict[str, Any]) -> list[dict[str, str]]:
    sources: list[dict[str, str]] = []

    def visit(value: Any) -> None:
        if isinstance(value, dict):
            if value.get("type") == "url_citation" and value.get("url"):
                sources.append(
                    {
                        "title": str(value.get("title") or ""),
                        "url": str(value.get("url") or ""),
                    }
                )
            action_sources = value.get("sources")
            if isinstance(action_sources, list):
                for source in action_sources:
                    if isinstance(source, dict) and source.get("url"):
                        sources.append(
                            {
                                "title": str(source.get("title") or ""),
                                "url": str(source.get("url") or ""),
                            }
                        )
            for nested in value.values():
                visit(nested)
        elif isinstance(value, list):
            for nested in value:
                visit(nested)

    visit(response_dict)
    deduped: list[dict[str, str]] = []
    seen: set[str] = set()
    for source in sources:
        url = source["url"]
        if url and url not in seen:
            seen.add(url)
            deduped.append(source)
    return deduped


def run_openai_search(item: SearchItem, model: str) -> dict[str, Any]:
    from openai import OpenAI

    client = OpenAI()
    prompt = f"""
다음 기관/분류에 대한 실제 인터넷 검색을 수행하고, 활용사례 후보를 한국어로 요약하세요.

검색 질의: {item.query}

출력 형식:
- 핵심 요약 3줄 이내
- 확인된 활용사례 후보
- 출처 URL

주의:
- 출처로 확인되지 않는 내용은 추정이라고 표시하세요.
- 기관명이 유사한 경우 동일 기관인지 조심해서 판단하세요.
""".strip()

    response = client.responses.create(
        model=model,
        tools=[{"type": "web_search"}],
        include=["web_search_call.action.sources"],
        input=prompt,
    )
    response_dict = response.model_dump(mode="json")
    return {
        "status": "success",
        "summary": response.output_text,
        "sources": extract_sources(response_dict),
        "raw_response_id": response_dict.get("id"),
    }


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "gpt_search_results"
    headers = ["category", "sheet", "row_number", "institution", "query", "status", "summary", "sources", "error"]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row.get("category"),
                row.get("sheet"),
                row.get("row_number"),
                row.get("institution"),
                row.get("query"),
                row.get("status"),
                row.get("summary"),
                json.dumps(row.get("sources", []), ensure_ascii=False),
                row.get("error"),
            ]
        )
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 80)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run GPT web searches for workbook rows.")
    parser.add_argument("--project-dir", default=".")
    parser.add_argument("--input-file", required=True)
    parser.add_argument("--mode", choices=["dry-run", "execute"], default=os.environ.get("GPT_SEARCH_MODE", "dry-run"))
    parser.add_argument("--max-rows", type=int, default=int(os.environ.get("GPT_SEARCH_MAX_ROWS", "10")))
    parser.add_argument("--model", default=os.environ.get("OPENAI_MODEL", "gpt-5.5"))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    project_dir = Path(args.project_dir).resolve()
    load_project_env(project_dir)

    input_path = resolve_input(project_dir, args.input_file)
    items = collect_search_items(input_path, args.max_rows)
    run_date = datetime.now().strftime("%Y%m%d")
    output_dir = project_dir / "outputs" / "gpt_search" / run_date
    output_dir.mkdir(parents=True, exist_ok=True)

    results: list[dict[str, Any]] = []
    api_key = os.environ.get("OPENAI_API_KEY")
    execute = args.mode == "execute"
    if execute and not api_key:
        raise RuntimeError("OPENAI_API_KEY is required for --mode execute.")

    for item in items:
        base = {
            "category": item.category,
            "sheet": item.sheet,
            "row_number": item.row_number,
            "institution": item.institution,
            "query": item.query,
            "row": item.row,
        }
        if not execute:
            results.append({**base, "status": "planned", "summary": "", "sources": []})
            continue
        try:
            search_result = run_openai_search(item, args.model)
            results.append({**base, **search_result})
        except Exception as exc:
            results.append({**base, "status": "failed", "summary": "", "sources": [], "error": str(exc)})

    stem = safe_filename(input_path.stem)
    jsonl_path = output_dir / f"{stem}_gpt_search.jsonl"
    xlsx_path = output_dir / f"{stem}_gpt_search.xlsx"
    write_jsonl(jsonl_path, results)
    write_xlsx(xlsx_path, results)

    payload = {
        "status": "success",
        "mode": args.mode,
        "input_file": str(input_path),
        "item_count": len(results),
        "jsonl_path": str(jsonl_path),
        "xlsx_path": str(xlsx_path),
        "sheets_detected": sorted({item.sheet for item in items}),
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    print(f"::{json.dumps({'outputs': payload}, ensure_ascii=False)}::")


if __name__ == "__main__":
    main()
